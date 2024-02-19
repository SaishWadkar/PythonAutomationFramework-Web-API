import openpyxl
'''
    4) basic operations on excel file (common activities) :
    1. get row count of sheet
    2. get column count of sheet
    3. read from cell
    4. write to cell
'''

class ExcelUtils:

    # to set workbook and sheet name
    def __init__(self,file,sheet):
        self.work_book = openpyxl.load_workbook(filename=file)
        # self.sheet = self.work_book.get_sheet_by_name(sheet)
        self.sheet = self.work_book[sheet]
        self.wb = file

    def get_row_count(self):
        return self.sheet.max_row

    def get_column_count(self):
        return self.sheet.max_column

    def read_cell_data(self,row_no,col_no):
        return self.sheet.cell(row=row_no,column=col_no).value


    def write_cell_data(self,row_no,col_no,data):
        self.sheet.cell(row_no,col_no).value = data
        self.work_book.save(self.wb)

        # self.sheet.cell(row=row_no, column=col_no).value = data
        # self.work_book.save(self.work_book)


    # reading entire row and storing it in form of list for better ussage
    def read_sheet(self):
        data = []
        for row in self.sheet.iter_rows(values_only=True):
            data.append(row)

        return data
